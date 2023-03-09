import os
import glob
import openpyxl
import csv

from openpyxl import load_workbook
from datetime import datetime

# Set the path to the folder containing the Excel workbooks
path = "C:/Heather's format - Copy"
new_folder_path = "C:/Weights test"

# Set the name of the sheet you want to save as a CSV
sheet_name = "hj_weights"

# Loop through all of the Excel workbooks in the folder
for file in glob.glob(os.path.join(path, "*.xlsx")):

    # Load the workbook
    wb = openpyxl.load_workbook(filename=file, data_only=True)

    # Get the specified sheet
    ws = wb[sheet_name]

    # Get the original workbook name
    workbook_name = os.path.basename(file)

    # Set the name for the CSV file using the custom name dictionary
    csv_name = workbook_name.replace(".xlsx", "") + "_" + sheet_name[3:] + ".csv"

    # Set the path for the CSV file
    csv_path = os.path.join(new_folder_path, csv_name)

    # Save the sheet as a CSV file
    with open(csv_path, 'w', newline='') as f:
        output = csv.writer(f, delimiter=",")
        for row in ws.iter_rows():
            output_row = []
            for cell in row:
                if isinstance(cell.value, datetime):
                    output_row.append(cell.value.date())
                else:
                    output_row.append(cell.value)
            output.writerow(output_row)